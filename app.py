import streamlit as st
import numpy as np
import pandas as pd
import os

# Importação do módulo de login independente
from login import renderizar_login

# ==========================================
# CONFIGURAÇÕES DA PÁGINA
# ==========================================
st.set_page_config(
    page_title="Controle de Patrimônio - GTI-SESA",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==========================================
# CSS CUSTOMIZADO PARA AUTOAJUSTE EM CELULARES
# ==========================================
st.markdown("""
    <style>
        /* Ajustes Globais para Telas Pequenas (Mobile) */
        @media (max-width: 768px) {
            .main .block-container {
                padding-left: 0.6rem !important;
                padding-right: 0.6rem !important;
                padding-top: 1rem !important;
                padding-bottom: 2rem !important;
            }
            h1 {
                font-size: 1.5rem !important;
                text-align: center;
            }
            h2 {
                font-size: 1.2rem !important;
            }
            h3 {
                font-size: 1.05rem !important;
            }
            /* Previne zoom automático do Safari no iOS ao focar em inputs */
            input, select, textarea {
                font-size: 16px !important;
            }
            /* Botões grandes e fáceis de tocar com o polegar */
            .stButton > button, .stDownloadButton > button {
                width: 100% !important;
                min-height: 48px !important;
                font-size: 16px !important;
                font-weight: bold !important;
                border-radius: 8px !important;
            }
            /* Ajuste visual das abas */
            .stTabs [data-baseweb="tab-list"] {
                gap: 2px !important;
            }
            .stTabs [data-baseweb="tab"] {
                padding: 8px 12px !important;
                font-size: 13px !important;
            }
        }

        /* Container do Scanner de Câmera Fluido */
        .scanner-wrapper {
            width: 100%;
            max-width: 100%;
            margin: auto;
        }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# VERIFICAÇÃO DE AUTENTICAÇÃO E LOGIN
# ==========================================
if not renderizar_login():
    st.stop()

if "pagina_atual" not in st.session_state:
    st.session_state.pagina_atual = "portal"

# ==========================================
# BARRA LATERAL (MENU E LOGOUT)
# ==========================================
with st.sidebar:
    st.markdown("### 👤 Usuário Autenticado")
    
    if st.session_state.pagina_atual == "inventario":
        if st.button("🏠 Voltar ao Portal"):
            st.session_state.pagina_atual = "portal"
            st.rerun()

    if st.button("🚪 Sair do Sistema"):
        st.session_state.autenticado = False
        st.session_state.pagina_atual = "portal"
        st.rerun()

# ==========================================
# TELA INTERMEDIÁRIA (PORTAL DE SERVIÇOS)
# ==========================================
if st.session_state.pagina_atual == "portal":
    st.title("🖥️ Portal de Sistemas GTI-SESA")
    st.markdown("Bem-vindo ao painel central de aplicações. Escolha o sistema que deseja acessar:")
    st.divider()

    col1, col2, col3 = st.columns([1, 2, 1])

    lista_urs = [
        "Selecione uma URS...",
        "URS Boa Vista", "URS Feu Rosa", "URS Jacaraípe", 
        "URS Novo Horizonte", "URS Serra Sede", "URS Serra Dourada"
    ]

    lista_ubs = [
        "Selecione uma UBS...",
        "UBS André Carloni", "UBS Bairro de Fátima", "UBS Feu Rosa", "UBS Barcelona", 
        "UBS Barro Branco", "UBS Campinho da Serra", "UBS Carapebus", "UBS Carapina Grande", 
        "UBS Central Carapina", "UBS Cidade Continental", "UBS Eldorado", "UBS Jardim Carapina", 
        "UBS Jardim Tropical", "UBS José de Anchieta", "UBS Laranjeiras Velha", "UBS Manguinhos", 
        "UBS Manoel Plaza", "UBS Nova Almeida", "UBS Nova Carapina I", "UBS Nova Carapina II", 
        "UBS Oceania", "UBS Pitanga", "UBS Planalto Serrano (Bloco A)", "UBS Planalto Serrano (Bloco B)", 
        "UBS Porto Canoa", "UBS São Diogo", "UBS São Marcos", "UBS Taquara I", "UBS Taquara II", 
        "UBS Vila Nova de Colares", "UBS Vista da Serra", "UBS Itinerante (atendimento na área rural)"
    ]

    with col2:
        # --- 1. MÓDULO DE INVENTÁRIO ---
        with st.container(border=True):
            st.markdown("<h3 style='text-align: center;'>📦 Módulo de Inventário</h3>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #666;'>Acesse a ferramenta de gestão, leitura de códigos de barra e controle de patrimônio.</p>", unsafe_allow_html=True)
            st.write("")

            urs_selecionada = st.selectbox("URS - Unidade Regional de Saúde", lista_urs, key="sel_urs")
            ubs_selecionada = st.selectbox("UBS - Unidade Básica de Saúde", lista_ubs, key="sel_ubs")

            if urs_selecionada != "Selecione uma URS...":
                st.session_state.saved_setor = urs_selecionada
            elif ubs_selecionada != "Selecione uma UBS...":
                st.session_state.saved_setor = ubs_selecionada

            st.write("")
            
            if st.button("📂 Abrir Inventário nesta Aba", use_container_width=True, type="primary", key="btn_inventario"):
                st.session_state.pagina_atual = "inventario"
                st.rerun()

        st.write("")

        # --- 2. ENTRADA DE EQUIPAMENTOS ---
        with st.container(border=True):
            st.markdown("<h3 style='text-align: center;'>📥 Entrada de Equipamentos</h3>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #666;'>Acesse a ferramenta de registro e recebimento de equipamentos nas unidades.</p>", unsafe_allow_html=True)
            st.write("")

            urs_entrada = st.selectbox("URS - Unidade Regional de Saúde", lista_urs, key="sel_urs_entrada")
            ubs_entrada = st.selectbox("UBS - Unidade Básica de Saúde", lista_ubs, key="sel_ubs_entrada")

            if urs_entrada != "Selecione uma URS...":
                st.session_state.saved_setor = urs_entrada
            elif ubs_entrada != "Selecione uma UBS...":
                st.session_state.saved_setor = ubs_entrada

            st.write("")
            
            if st.button("📂 Abrir Entrada nesta Aba", use_container_width=True, type="primary", key="btn_entrada"):
                st.session_state.pagina_atual = "inventario"
                st.rerun()

        st.write("")

        # --- 3. SAÍDA DE EQUIPAMENTOS ---
        with st.container(border=True):
            st.markdown("<h3 style='text-align: center;'>📤 Saída de Equipamentos</h3>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #666;'>Acesse a ferramenta de baixa, transferência e saída de equipamentos.</p>", unsafe_allow_html=True)
            st.write("")

            urs_saida = st.selectbox("URS - Unidade Regional de Saúde", lista_urs, key="sel_urs_saida")
            ubs_saida = st.selectbox("UBS - Unidade Básica de Saúde", lista_ubs, key="sel_ubs_saida")

            if urs_saida != "Selecione uma URS...":
                st.session_state.saved_setor = urs_saida
            elif ubs_saida != "Selecione uma UBS...":
                st.session_state.saved_setor = ubs_saida

            st.write("")
            
            if st.button("📂 Abrir Saída nesta Aba", use_container_width=True, type="primary", key="btn_saida"):
                st.session_state.pagina_atual = "inventario"
                st.rerun()

    st.stop()

# ==========================================
# CONFIGURAÇÕES E CONSTANTES
# ==========================================
ARQUIVO_EXCEL = "Tabela_Patrimonios_UBS_Feu_Rosa.xlsx"
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/12mNKTWLExRwZx3EKSB78oTScQk6ctGvi6eNKt5QyXEw/edit?usp=sharing"

COLUNAS_PADRAO = ["Local / Setor", "Patrimônio PC", "Patrimônio Tela", "Patrimônio Nobreak"]

@st.cache_resource
def obter_conexao_gsheets():
    try:
        from streamlit_gsheets import GSheetsConnection
        return st.connection("gsheets", type=GSheetsConnection)
    except Exception:
        return None

conn = obter_conexao_gsheets()

# ==========================================
# FUNÇÕES DE MANIPULAÇÃO DE DADOS
# ==========================================
def padronizar_e_organizar_df(df: pd.DataFrame) -> pd.DataFrame:
    for col in COLUNAS_PADRAO:
        if col not in df.columns:
            df[col] = ""

    outras_colunas = [c for c in df.columns if c not in COLUNAS_PADRAO]
    ordem_final = COLUNAS_PADRAO + outras_colunas

    return df[ordem_final].fillna("").astype(str)

def aplicar_estilo_excel(caminho_arquivo: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb['Patrimônios']

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    row_fill_even = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 22
        fill = row_fill_even if r % 2 == 0 else row_fill_odd
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = align_left if c == 1 else align_center

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    wb.save(caminho_arquivo)

@st.cache_data(ttl=60)
def carregar_dados_excel() -> tuple[pd.DataFrame, list[str]]:
    if os.path.exists(ARQUIVO_EXCEL):
        try:
            df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Patrimônios', dtype=str, keep_default_na=False)
            
            if not df.empty and df.columns[0].startswith("Tabela de Patrimônios"):
                df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Patrimônios', header=1, dtype=str, keep_default_na=False)

            df = df.dropna(how='all')
            df = padronizar_e_organizar_df(df)
            return df, list(df.columns)
        except Exception as e:
            st.error(f"Erro ao carregar a planilha existente: {e}")
    
    df_empty = pd.DataFrame(columns=COLUNAS_PADRAO)
    df_empty = padronizar_e_organizar_df(df_empty)
    return df_empty, COLUNAS_PADRAO

def salvar_no_excel(df: pd.DataFrame) -> None:
    df = padronizar_e_organizar_df(df)
    
    try:
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Patrimônios', index=False)
        aplicar_estilo_excel(ARQUIVO_EXCEL)
        carregar_dados_excel.clear()
    except Exception as e:
        st.error(f"Erro ao salvar e formatar a planilha local: {e}")

    if conn is not None:
        try:
            conn.update(
                spreadsheet=GOOGLE_SHEET_URL,
                data=df
            )
            st.toast("☁️ Dados sincronizados no Google Sheets!")
        except Exception as e:
            st.error(f"Falha na sincronização com Google Sheets: {e}")

def adicionar_e_salvar(codigo: str, descricao: str, setor: str) -> None:
    df, _ = carregar_dados_excel()
    
    coluna_alvo = descricao.strip()
    setor_limpo = setor.strip() if setor else "Não informado"
    codigo_limpo = str(codigo).strip()
    
    if coluna_alvo not in df.columns:
        df[coluna_alvo] = ""
    
    df = padronizar_e_organizar_df(df)
    
    df["Local / Setor"] = df["Local / Setor"].str.strip()
    mascara_setor = df["Local / Setor"].str.lower() == setor_limpo.lower()
    
    if mascara_setor.any():
        idx = df[mascara_setor].index[0]
        df.at[idx, coluna_alvo] = codigo_limpo
    else:
        nova_linha = {col: "" for col in df.columns}
        nova_linha["Local / Setor"] = setor_limpo
        nova_linha[coluna_alvo] = codigo_limpo
        
        df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)
    
    df = padronizar_e_organizar_df(df)
    salvar_no_excel(df)
    st.session_state.df_historico = df

def processar_imagem(image_file) -> tuple:
    import cv2
    import zxingcpp
    try:
        if hasattr(image_file, 'getvalue'):
            file_bytes = np.frombuffer(image_file.getvalue(), dtype=np.uint8)
        else:
            file_bytes = np.asarray(bytearray(image_file.read()), dtype=np.uint8)
        
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None:
            return None, []

        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        barcodes = zxingcpp.read_barcodes(img_rgb)
        resultados = []

        for barcode in barcodes:
            if hasattr(barcode, 'position') and barcode.position:
                try:
                    pos = barcode.position
                    if hasattr(pos, 'top_left'):
                        pts_list = [
                            [int(pos.top_left.x), int(pos.top_left.y)],
                            [int(pos.top_right.x), int(pos.top_right.y)],
                            [int(pos.bottom_right.x), int(pos.bottom_right.y)],
                            [int(pos.bottom_left.x), int(pos.bottom_left.y)]
                        ]
                    else:
                        pts_list = [[int(getattr(pt, 'x', pt[0])), int(getattr(pt, 'y', pt[1]))] for pt in pos]

                    if pts_list:
                        pts = np.array(pts_list, np.int32).reshape((-1, 1, 2))
                        cv2.polylines(img_rgb, [pts], True, (0, 255, 0), 3)
                except Exception:
                    pass

            resultados.append({
                "codigo": barcode.text, 
                "tipo": str(barcode.format).replace("BarcodeFormat.", "")
            })

        return img_rgb, resultados
    except Exception as e:
        st.error(f"Erro ao processar imagem: {e}")
        return None, []

# ==========================================
# INICIALIZAÇÃO DE ESTADO
# ==========================================
df_inicial, colunas_iniciais = carregar_dados_excel()

if "df_historico" not in st.session_state:
    st.session_state.df_historico = df_inicial

if "saved_setor" not in st.session_state:
    st.session_state.saved_setor = ""

if "saved_descricao" not in st.session_state:
    st.session_state.saved_descricao = ""

# ==========================================
# INTERFACE PRINCIPAL
# ==========================================
st.title("📦 Controle de Patrimônio GTI-SESA")
st.divider()

st.subheader("1. Selecione o Local e a Coluna")

opcoes_patrimonio = [col for col in st.session_state.df_historico.columns if col != "Local / Setor"]
opcoes_patrimonio.append("➕ Outra descrição (Criar nova coluna ao final)")

col_desc1, col_desc2, col_desc3 = st.columns([1, 1, 1])

with col_desc1:
    setor_input = st.text_input("Local / Setor:", value=st.session_state.saved_setor, placeholder="Ex: Consultório 1...", key="setor_input_key")
    st.session_state.saved_setor = setor_input

with col_desc2:
    opcao_selecionada = st.selectbox(
        "Coluna de destino:",
        opcoes_patrimonio,
        key="opcao_selecionada_key"
    )

with col_desc3:
    if opcao_selecionada == "➕ Outra descrição (Criar nova coluna ao final)":
        descricao_final = st.text_input("Nome da nova coluna:", placeholder="Ex: Patrimônio Impressora", key="descricao_nova_key")
    else:
        descricao_final = opcao_selecionada
    
    st.session_state.saved_descricao = descricao_final

st.divider()

st.subheader("2. Realize a Leitura do Código")

if not descricao_final or not setor_input.strip():
    st.warning("⚠️ Preencha o **Local / Setor** e selecione a **Coluna de Destino** para ativar o leitor.")
else:
    tab_unificada, tab_upload = st.tabs([
        "⚡ Câmera do Celular / Scanner USB", 
        "📁 Upload de Imagem"
    ])

    with tab_unificada:
        st.markdown(f"📍 Setor: **`{setor_input}`** | Coluna: **`{descricao_final}`**")
        
        col_camera, col_usb = st.columns([1.2, 1])

        with col_camera:
            st.markdown("##### 📱 Câmera (Bip e Registro Automático)")
            st.caption("Aponte a câmera para o código de barras. A leitura e salvamento são instantâneos.")

            html_scanner = """
            <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
            <script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>

            <style>
                #reader {
                    width: 100% !important;
                    max-width: 100% !important;
                    border-radius: 12px;
                    overflow: hidden;
                    border: 2px solid #1F4E78;
                    background-color: #000;
                }
                #reader video {
                    object-fit: cover !important;
                    border-radius: 10px;
                }
                #scan-status {
                    text-align: center;
                    margin-top: 8px;
                    font-weight: bold;
                    color: #1F4E78;
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                    font-size: 14px;
                }
            </style>

            <div class="scanner-wrapper">
                <div id="reader"></div>
                <div id="scan-status">📷 Inicializando câmera...</div>
            </div>

            <script>
                let lastScannedCode = "";
                let lastScannedTime = 0;
                let audioCtx = null;

                function tocarBipNativo() {
                    try {
                        if (!audioCtx) {
                            audioCtx = new (window.AudioContext || window.webkitAudioContext)();
                        }
                        if (audioCtx.state === 'suspended') {
                            audioCtx.resume();
                        }
                        const osc = audioCtx.createOscillator();
                        const gain = audioCtx.createGain();
                        osc.type = 'sine';
                        osc.frequency.setValueAtTime(1500, audioCtx.currentTime);
                        gain.gain.setValueAtTime(0.4, audioCtx.currentTime);
                        osc.connect(gain);
                        gain.connect(audioCtx.destination);
                        osc.start();
                        osc.stop(audioCtx.currentTime + 0.12);
                    } catch(e) { console.log(e); }
                }

                function onScanSuccess(decodedText, decodedResult) {
                    const agora = Date.now();
                    if (decodedText === lastScannedCode && (agora - lastScannedTime) < 3000) {
                        return; 
                    }

                    lastScannedCode = decodedText;
                    lastScannedTime = agora;

                    tocarBipNativo();
                    document.getElementById('scan-status').innerText = "✅ Lido: " + decodedText + " (Salvando...)";

                    const parentDoc = window.parent.document;
                    const inputEl = parentDoc.querySelector('input[placeholder*="Aguardando bipagem"]');
                    
                    if (inputEl) {
                        const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, "value").set;
                        nativeSetter.call(inputEl, decodedText);
                        inputEl.dispatchEvent(new Event('input', { bubbles: true }));

                        setTimeout(() => {
                            const buttons = Array.from(parentDoc.querySelectorAll('button'));
                            const submitBtn = buttons.find(b => b.innerText.includes('Registrar Manualmente'));
                            
                            if (submitBtn) {
                                submitBtn.click();
                            } else {
                                inputEl.dispatchEvent(new KeyboardEvent('keydown', {
                                    bubbles: true, cancelable: true, key: 'Enter', code: 'Enter', keyCode: 13
                                }));
                            }
                            
                            setTimeout(() => {
                                document.getElementById('scan-status').innerText = "📷 Próximo código...";
                            }, 1500);
                            
                        }, 150);
                    }
                }

                function getQrBoxDimensions(viewfinderWidth, viewfinderHeight) {
                    const minEdge = Math.min(viewfinderWidth, viewfinderHeight);
                    const width = Math.floor(minEdge * 0.85);
                    const height = Math.floor(width * 0.55);
                    return { width: Math.max(width, 200), height: Math.max(height, 100) };
                }

                const html5QrCode = new Html5Qrcode("reader");
                const config = { 
                    fps: 25, 
                    qrbox: getQrBoxDimensions,
                    experimentalFeatures: { useBarCodeDetectorIfSupported: true }
                };

                html5QrCode.start(
                    { facingMode: "environment" }, 
                    config, 
                    onScanSuccess
                ).then(() => {
                    document.getElementById('scan-status').innerText = "📷 Leitor pronto. Aponte para o código.";
                }).catch(err => {
                    html5QrCode.start({ facingMode: "user" }, config, onScanSuccess);
                });
            </script>
            """
            st.components.v1.html(html_scanner, height=390)

        with col_usb:
            st.markdown("##### 🔌 Entrada Manual / Scanner USB")
            with st.form(key="form_bipagem", clear_on_submit=True):
                codigo_input = st.text_input(
                    "Código Lido / Bipado:", 
                    autocomplete="off",
                    placeholder="Aguardando bipagem...",
                    key="input_codigo_bip"
                )
                btn_adicionar = st.form_submit_button("Registrar Manualmente", type="primary", use_container_width=True)

                if btn_adicionar and codigo_input.strip():
                    adicionar_e_salvar(codigo_input.strip(), descricao_final, setor_input)
                    st.success(f"✅ Registrado: `{codigo_input.strip()}` em **'{descricao_final}'**")
                    st.rerun()

    with tab_upload:
        st.markdown(f"📍 Setor: **`{setor_input}`** | Coluna: **`{descricao_final}`**")
        uploaded_file = st.file_uploader("Escolha uma imagem contendo o código", type=["jpg", "png", "jpeg"])

        if uploaded_file is not None:
            img_processada, codigos_encontrados = processar_imagem(uploaded_file)
            
            col_img1, col_img2 = st.columns(2)
            with col_img1:
                if img_processada is not None:
                    st.image(img_processada, caption="Imagem Processada", use_container_width=True)
                
            with col_img2:
                if codigos_encontrados:
                    st.success(f"{len(codigos_encontrados)} código(s) detectado(s)!")
                    for item in codigos_encontrados:
                        adicionar_e_salvar(item['codigo'], descricao_final, setor_input)
                        st.write(f"**Código:** `{item['codigo']}` ➡️ Coluna: **{descricao_final}**")
                    st.rerun()
                else:
                    st.error("Nenhum código de barras identificado na imagem.")

st.divider()
st.header("📊 Tabela de Patrimônios")

df_atual, _ = carregar_dados_excel()
if not df_atual.empty:
    df_exibicao = df_atual.fillna("").astype(str)
    st.dataframe(df_exibicao, use_container_width=True)

    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        if st.button("🔄 Recarregar Planilha", use_container_width=True):
            carregar_dados_excel.clear()
            st.rerun()
            
    with col_btn2:
        csv = df_atual.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="⬇️ Baixar Tabela (CSV)",
            data=csv,
            file_name="Tabela_Patrimonios_UBS_Feu_Rosa.csv",
            mime="text/csv",
            use_container_width=True
        )
else:
    st.info("A planilha está sem registros até o momento.")